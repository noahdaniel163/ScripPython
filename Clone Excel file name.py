from openpyxl import load_workbook
from shutil import copyfile
from datetime import datetime, timedelta
import os

# Path to the original Excel file
original_file_path = r'E:\1\C00334-79660001-79660001-yyyymm-ST-N-01.xlsx'

# Destination folder for cloned files
destination_folder = r'E:\1\Cloned_Files'

# Ensure the destination folder exists
os.makedirs(destination_folder, exist_ok=True)

# Load the original Excel file
wb = load_workbook(original_file_path)
ws = wb.active

# Define the start and end dates for cloning (September 2023 to February 2024)
start_date = datetime(2023, 9, 1)
end_date = datetime(2024, 3, 1)

# Loop through each month from start_date to end_date
current_date = start_date
while current_date < end_date:
    # Generate the new file name based on the current month
    new_file_name = f"C00334-79660001-79660001-{current_date.strftime('%Y%m')}-ST-N-01.xlsx"

    # Clone the original file
    copyfile(original_file_path, os.path.join(destination_folder, new_file_name))

    # Update the specified cells in the cloned file
    cloned_wb = load_workbook(os.path.join(destination_folder, new_file_name))
    cloned_ws = cloned_wb.active
    cloned_ws['C2'] = new_file_name
    cloned_ws['C5'] = current_date.strftime('%Y%m')
    cloned_wb.save(os.path.join(destination_folder, new_file_name))

    # Move to the next month
    current_date = current_date.replace(day=1) + timedelta(days=32)
    current_date = current_date.replace(day=1)

print("Cloning and updating cells completed.")
