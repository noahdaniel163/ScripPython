import sys
import os
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableView, QDialog, QGridLayout, QLabel, QLineEdit, QPushButton
from PyQt5.QtCore import Qt, QAbstractTableModel, QVariant, QModelIndex

class AcronymModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, parent=QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        return len(self._data[0]) if self._data else 0

    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            row = index.row()
            col = index.column()
            value = self._data[row][col]
            return QVariant(value)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return QVariant(['Acronym', 'Explanation'][section])
            elif orientation == Qt.Vertical:
                return QVariant(str(section + 1))

    def setData(self, index, value, role=Qt.EditRole):
        if role == Qt.EditRole:
            row = index.row()
            col = index.column()
            self._data[row][col] = value
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
            return True
        return False

    def insertRows(self, position, rows, parent=QModelIndex()):
        self.beginInsertRows(parent, position, position + rows - 1)
        for i in range(rows):
            self._data.insert(position + i, ['', ''])
        self.endInsertRows()
        return True

    def removeRows(self, position, rows, parent=QModelIndex()):
        self.beginRemoveRows(parent, position, position + rows - 1)
        del self._data[position:position+rows]
        self.endRemoveRows()
        return True

class AddDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Add Acronym')
        self.resize(300, 100)
        layout = QGridLayout()

        self.acronym_label = QLabel('Acronym:')
        self.acronym_edit = QLineEdit()
        self.explanation_label = QLabel('Explanation:')
        self.explanation_edit = QLineEdit()
        self.ok_button = QPushButton('OK')
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton('Cancel')
        self.cancel_button.clicked.connect(self.reject)

        layout.addWidget(self.acronym_label, 0, 0)
        layout.addWidget(self.acronym_edit, 0, 1)
        layout.addWidget(self.explanation_label, 1, 0)
        layout.addWidget(self.explanation_edit, 1, 1)
        layout.addWidget(self.ok_button, 2, 0)
        layout.addWidget(self.cancel_button, 2, 1)

        self.setLayout(layout)

    def get_data(self):
        return (self.acronym_edit.text(), self.explanation_edit.text())

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Acronym Table')
        self.resize(500, 300)
        self.table_view = QTableView()
        self.setCentralWidget(self.table_view)
        self.acronym_model = AcronymModel([])
        self.table_view.setModel(self.acronym_model)

        self.add_button = QPushButton('Add')
        self.add_button.clicked.connect(self.add_data)
        self.remove_button = QPushButton('Remove')
self.remove_button.clicked.connect(self.remove_data)

        toolbar = self.addToolBar('Toolbar')
        toolbar.addWidget(self.add_button)
        toolbar.addWidget(self.remove_button)

    def add_data(self):
        dialog = AddDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            acronym, explanation = dialog.get_data()
            self.acronym_model.insertRow(len(self.acronym_model._data))
            row = len(self.acronym_model._data) - 1
            self.acronym_model.setData(self.acronym_model.index(row, 0), acronym)
            self.acronym_model.setData(self.acronym_model.index(row, 1), explanation)

    def remove_data(self):
        selection = self.table_view.selectedIndexes()
        rows = set()
        for index in selection:
            rows.add(index.row())
        for row in reversed(sorted(rows)):
            self.acronym_model.removeRow(row)

    def save_data(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Acronym', 'Explanation'])
        for row in self.acronym_model._data:
            ws.append(row)
        filename, _ = QFileDialog.getSaveFileName(self, 'Save File', os.getenv('HOME'), 'Excel files (*.xlsx)')
        if filename:
            wb.save(filename)

    def load_data(self):
        filename, _ = QFileDialog.getOpenFileName(self, 'Open File', os.getenv('HOME'), 'Excel files (*.xlsx)')
        if filename:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(list(row))
            self.acronym_model = AcronymModel(data)
            self.table_view.setModel(self.acronym_model)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
