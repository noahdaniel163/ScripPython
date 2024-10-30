import os
from openpyxl import load_workbook, Workbook

# Đường dẫn đến thư mục chứa các file excel
folder_path = r'E:\2022\2022_Edit'

# Thứ tự đúng của các file theo tháng
sorted_files = [
    'Jan2022.xlsx', 'Feb2022.xlsx', 'Mar2022.xlsx', 'Apr2022.xlsx', 'May2022.xlsx',
    'June2022.xlsx', 'July2022.xlsx', 'Aug2022.xlsx', 'Sep2022.xlsx', 'Oct2022.xlsx',
    'Nov2022.xlsx', 'Dec2022.xlsx'
]

# Tạo một workbook mới để chứa các sheet được ghép
output_wb = Workbook()

# Đọc và ghép dữ liệu từ các file đã sắp xếp
for filename in sorted_files:
    file_path = os.path.join(folder_path, filename)
    
    # Mở workbook của file hiện tại
    try:
        wb = load_workbook(file_path)
        
        # Lấy sheet thứ 3 (index 2)
        sheet = wb.worksheets[2]  # Sheet thứ 3 trong file
        
        # Tạo một sheet mới trong workbook kết quả và sao chép nội dung
        new_sheet = output_wb.create_sheet(title=sheet.title[:31])  # Giới hạn tên sheet tối đa 31 ký tự
        
        for row in sheet.iter_rows():
            for cell in row:
                # Sao chép giá trị của ô (không sao chép toàn bộ định dạng để tránh lỗi)
                new_sheet[cell.coordinate].value = cell.value
                
                # Nếu muốn sao chép một số định dạng cơ bản như định dạng số
                new_sheet[cell.coordinate].number_format = cell.number_format

    except Exception as e:
        print(f"Lỗi khi đọc file {filename}: {e}")

# Xóa sheet mặc định được tạo tự động
if "Sheet" in output_wb.sheetnames:
    output_wb.remove(output_wb["Sheet"])

# Lưu workbook kết quả vào file mới
output_file = r'E:\2022\2022_Edit\2022.xlsx'
output_wb.save(output_file)

print(f"Đã lưu file Excel hợp nhất tại {output_file}")
