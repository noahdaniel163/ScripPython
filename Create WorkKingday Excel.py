from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

# Tạo workbook mới
wb = Workbook()

# Lấy active sheet (sheet mặc định)
ws = wb.active

# Xóa sheet mặc định
wb.remove(ws)

# Tạo sheet cho từng tháng từ tháng 8 đến tháng 12 của năm 2023
for month in range(8, 13):
    # Xác định ngày đầu tiên của tháng
    dt = datetime(2023, month, 1)
    
    # Kiểm tra nếu tháng hiện tại không phải là thứ 7 hoặc chủ nhật
    if dt.weekday() != 5 and dt.weekday() != 6:
        # Tạo tên sheet theo định dạng "mmm-2023"
        sheet_name = dt.strftime("%b-%Y")
        
        # Kiểm tra xem sheet đã tồn tại chưa
        if sheet_name not in wb.sheetnames:
            # Tạo sheet mới
            ws = wb.create_sheet(title=sheet_name)
            
            # Đặt tiêu đề cho cột A là "Ngày làm việc"
            ws["A1"].value = "Ngày làm việc"
            
            # Căn giữa tiêu đề trong cột A
            ws["A1"].alignment = Alignment(horizontal="center")
            
            # Đặt format ngày trong cột A là text
            ws["A1"].number_format = "@"
            
            # Xác định ngày đầu tiên của sheet
            start_date = dt
            
            # Tạo ngày tiếp theo để bắt đầu vòng lặp
            next_date = start_date + timedelta(days=1)
            
            # Đặt ngày làm việc trong cột A, bắt đầu từ dòng thứ 3
            row = 3
            while next_date.month == start_date.month:
                if next_date.weekday() != 5 and next_date.weekday() != 6:
                    ws["A" + str(row)].value = next_date.strftime("%d/%m/%Y")
                    row += 1
                next_date += timedelta(days=1)

# Lưu workbook vào file
wb.save("workbook.xlsx")
